# Objednávací nástroj pre Excel (GUI)
# Autor: Tadeáš Galbavý
# Licencia: Creative Commons BY-NC 4.0 (https://creativecommons.org/licenses/by-nc/4.0/)
#
# Tento skript je súčasťou aplikácie na automatizáciu objednávania produktov z Excelu.
# Môžete ho používať a upravovať pre nekomerčné účely s uvedením autora.
# Komerčné použitie je bez súhlasu autora zakázané.

# Import knižnic pre prácu s dátami, Excelom, GUI a systémom

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, messagebox, ttk, Button, Label, Entry, StringVar, Checkbutton, BooleanVar, Frame, Toplevel
import os
import time
import threading
import shutil
import tempfile
import sys
from decimal import Decimal, getcontext
import math

# Nastavenie presnosti pre desatinné výpočty (Decimal)
getcontext().prec = 6

# Funkcia na extrakciu ikony pre GUI (podporuje aj .exe build cez PyInstaller)
def extract_icon_temp():
    if hasattr(sys, "_MEIPASS"):
        
        zdroj = os.path.join(sys._MEIPASS, "logo.ico")
    else:
        
        zdroj = os.path.join(os.path.dirname(__file__), "logo.ico")

    ciel = os.path.join(tempfile.gettempdir(), "temp_logo.ico")
    shutil.copyfile(zdroj, ciel)
    return ciel

# Vypočíta minimálne počet ks pre dosiahnutie cieľového koeficientu s toleranciou
def najdi_min_objednavku(skladom, objednavky, priemer, ciel_koef):
    skladom = Decimal(str(skladom))
    objednavky = Decimal(str(objednavky))
    priemer = Decimal(str(priemer)) if priemer != 0 else Decimal("0.1")
    ciel_koef = Decimal(str(ciel_koef))
    tolerancia = Decimal("0.05")  # 5 % tolerancia

    aktualna_zasoba = skladom - objednavky
    aktualny_koef = aktualna_zasoba / priemer

    # Tolerancia vo výpočte
    if aktualny_koef >= ciel_koef - tolerancia:
        return 0

    for objednat in range(0, 10000):
        aktualna_zasoba = skladom - objednavky + objednat
        aktualny_koef = aktualna_zasoba / priemer
        if aktualny_koef >= ciel_koef - tolerancia:
            return objednat

    return 10000

# Hlavný algoritmus na výpočet objednávky podľa typu produktu a koeficientu
def vypocitaj_objednavku(row, koef_bestseller, koef_bezny):
    skladom = row["Reálne skladom"]
    objednavky = row["Počet nevybavených objednávok"] + row["Počet neštandardných objednávok"]
    priemer = row["Štvrťročný priemer"] if row["Štvrťročný priemer"] != 0 else 0.1

    if row["Dopredaj"] == "Ano":
        # Objednaj presne na vykrytie objednávok – nič navyše
        return max(objednavky - skladom, 0)

    if row["Na objednávku"] == "Ano":
       
        return max(objednavky - skladom, 0)

    # Bestseller → objednaj s rezervou 1 ks navyše
    # Bestseller → výpočet podľa vlastného koeficientu
    if row["Bestseller"] == "Ano":
        koef = koef_bestseller
        return najdi_min_objednavku(skladom, objednavky, priemer, koef)

    # Bežný výpočet cez koeficient
    koef = koef_bezny
    return najdi_min_objednavku(skladom, objednavky, priemer, koef)

# Spracovanie Excelu pre režim "vykrytie objednávok"
def uprav_excel_old(subor_cesta, vystup_cesta, zahrnut_bestsellery, progress_callback=None):
    df = pd.read_excel(subor_cesta)
    df["Štvrťročný priemer"] = pd.to_numeric(df["Štvrťročný priemer"], errors='coerce').fillna(0)
    df["Štvrťročný priemer"] = df["Štvrťročný priemer"].replace(0, 0.1)
    if progress_callback: progress_callback(10)

    # Pretypovanie 
    df["Reálne skladom"] = pd.to_numeric(df["Reálne skladom"], errors='coerce').fillna(0)
    df["Počet nevybavených objednávok"] = pd.to_numeric(df["Počet nevybavených objednávok"], errors='coerce').fillna(0)
    df["Počet neštandardných objednávok"] = pd.to_numeric(df["Počet neštandardných objednávok"], errors='coerce').fillna(0)
    df["Štvrťročný priemer"] = df["Štvrťročný priemer"].replace([0, '0'], 0.1)

    # Bezpečný výpočet
    if zahrnut_bestsellery:
        df["Calculation"] = df.apply(lambda row: max(
            0,
            (
                float(row["Počet nevybavených objednávok"]) +
                float(row["Počet neštandardných objednávok"]) -
                float(row["Reálne skladom"])
            ) if row["Dopredaj"] == "Ano" else
            (
                float(row["Počet nevybavených objednávok"]) +
                float(row["Počet neštandardných objednávok"]) +
                2 - float(row["Reálne skladom"])
            ) if row["Bestseller"] == "Ano" else
            float(row["Počet nevybavených objednávok"]) +
            float(row["Počet neštandardných objednávok"]) -
            float(row["Reálne skladom"])
            ), axis=1)
    else:
        df["Calculation"] = df["Reálne skladom"] - (
            df["Počet nevybavených objednávok"] + df["Počet neštandardných objednávok"])
        df["Calculation"] = df["Calculation"].apply(lambda x: 0 if x >= 0 else x * -1)
        
    df.at[0, "Režim výpočtu"] = "Vykrytie objednávok"
    df.at[0, "Bestseller zahrnutý"] = "Áno" if zahrnut_bestsellery else "Nie"

    if progress_callback: progress_callback(30)
    df.to_excel(vystup_cesta, index=False)
    wb = load_workbook(vystup_cesta)
    ws = wb.active
    podfarbi_a_oramuj(ws)
    wb.save(vystup_cesta)
    if progress_callback: progress_callback(100)
    return "OK"

# Spracovanie Excelu pre režim "objednávanie na sklad"
def uprav_excel_new(subor_cesta, vystup_cesta, koef_bestseller, koef_bezny, progress_callback=None):
    df = pd.read_excel(subor_cesta)
    if progress_callback: progress_callback(10)
    df["Štvrťročný priemer"] = pd.to_numeric(df["Štvrťročný priemer"], errors='coerce').fillna(0)
    df["Štvrťročný priemer"] = df["Štvrťročný priemer"].replace(0, 0.1)
    df["Calculation"] = df.apply(lambda row: vypocitaj_objednavku(row, koef_bestseller, koef_bezny), axis=1)
    
    if progress_callback: progress_callback(30)
    df.to_excel(vystup_cesta, index=False)
    wb = load_workbook(vystup_cesta)
    ws = wb.active

    # Dynamické doplnenie vzorca pre koeficient
    skladom_col = objednavky1_col = objednavky2_col = priemer_col = calc_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Reálne skladom": skladom_col = get_column_letter(idx)
        elif cell.value == "Počet nevybavených objednávok": objednavky1_col = get_column_letter(idx)
        elif cell.value == "Počet neštandardných objednávok": objednavky2_col = get_column_letter(idx)
        elif cell.value == "Štvrťročný priemer": priemer_col = get_column_letter(idx)
        elif cell.value == "Calculation": calc_col = get_column_letter(idx)

    if skladom_col and objednavky1_col and objednavky2_col and priemer_col and calc_col:
        koef_col_idx = ws.max_column + 1
        ws.cell(row=1, column=koef_col_idx, value="Vypočítaný KOEF")
        for row in range(2, ws.max_row + 1):
            formula = f"=({skladom_col}{row}-{objednavky1_col}{row}-{objednavky2_col}{row}+{calc_col}{row})/{priemer_col}{row}"
            cell = ws.cell(row=row, column=koef_col_idx, value=formula)
            cell.number_format = "0.0"

    podfarbi_a_oramuj(ws)

    # Pridanie meta informácií na koniec
    last_col = ws.max_column + 1
    ws.insert_cols(last_col)
    ws.insert_cols(last_col)
    ws.insert_cols(last_col)

    meta_headers = ["Režim výpočtu", "Koef. Bestseller", "Koef. bežný"]
    meta_values = ["Objednávanie na sklad", koef_bestseller, koef_bezny]
    for i in range(3):
        ws.cell(row=1, column=last_col + i, value=meta_headers[i])
        ws.cell(row=2, column=last_col + i, value=meta_values[i])
    wb.save(vystup_cesta)
    if progress_callback: progress_callback(100)
    return "OK"

# Funkcia na podfarbenie a orámovanie bunkových hodnôt v Exceli
# Používa farebnú logiku podľa typu stĺpca a pridáva orámovanie pre všetky bunky
def podfarbi_a_oramuj(ws):
    fills = {
        "bledomodrá": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "olivová": PatternFill(start_color="808000", end_color="808000", fill_type="solid"),
        "oranžová": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "červená": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
    }
    farby_stlpcov = {
        "Ročný priemer": fills["bledomodrá"],
        "Polročný priemer": fills["bledomodrá"],
        "Štvrťročný priemer": fills["bledomodrá"],
        "Reálne skladom": fills["olivová"],
        "Počet nevybavených objednávok": fills["oranžová"],
        "Počet neštandardných objednávok": fills["oranžová"],
        "Calculation": fills["červená"],
    }
    for col in ws.iter_cols(min_row=1):
        header = col[0].column_letter + '1'
        stlpec_nazov = ws[header].value
        if stlpec_nazov in farby_stlpcov:
            for cell in col:
                # 💡 formátuj na 0.0 ak je číslo a ide o priemer
                if stlpec_nazov in ["Ročný priemer", "Polročný priemer", "Štvrťročný priemer"]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.0"
                if cell.value is not None and not (isinstance(cell.value, float) and math.isnan(cell.value)):
                    cell.fill = farby_stlpcov[stlpec_nazov]

    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

# Spustenie hlavného GUI okna aplikácie
# Umožňuje výber režimu, zadanie parametrov a spustenie spracovania Excel súboru
def spust_gui():
    def zobraz_formular_old():
        vyber_rezim_frame.pack_forget()
        formular_old.pack(pady=10)
        spatne_tlacidlo.pack(pady=10)

    def zobraz_formular_new():
        vyber_rezim_frame.pack_forget()
        formular_new.pack(pady=10)
        spatne_tlacidlo.pack(pady=10)

    def vyber_subor(zahrnut_bestsellery=None, koef_b=None, koef_n=None):
        subor = filedialog.askopenfilename(title="Vyber Excel súbor weirdo", filetypes=[("Excel súbory", "*.xlsx")])
        if not subor:
            return

        progress_bar["value"] = 0
        progress_label["text"] = "Spracovanie prebieha..."

        def update_progress(value):
            progress_bar["value"] = value
            root.update_idletasks()
            time.sleep(0.15)

        def spracuj():
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    vystup = tmp.name
                if zahrnut_bestsellery is not None:
                    uprav_excel_old(subor, vystup, zahrnut_bestsellery, update_progress)
                else:
                    uprav_excel_new(subor, vystup, koef_b, koef_n, update_progress)

                
                for widget in root.winfo_children():
                    widget.pack_forget()

                Label(root, text="✅ Hotovo! Chceš uložiť výstupný súbor?", bg="#C5D1A3", font=("Segoe UI", 11)).pack(pady=15)

                def uloz():
                    nova_cesta = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
                    if not nova_cesta:
                        return

                    try:
                        # Skúsime otvoriť cieľový súbor na zápis (ak je zamknutý, padne)
                        with open(nova_cesta, "a"):
                            pass
                        shutil.copyfile(vystup, nova_cesta)
                        messagebox.showinfo("Uložené", f"Súbor bol uložený ako:\n{nova_cesta}")
                        root.destroy()
                    except PermissionError:
                        messagebox.showerror("Chyba", "Súbor je pravdepodobne otvorený.\nZavri ho a skús znova (lol).")

                Button(root, text="Uložiť ako...", command=uloz,
                    bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=10).pack(pady=5)

                Button(root, text="Zatvoriť", command=root.destroy,
                    bg="#999999", fg="white", font=("Segoe UI", 9), padx=8).pack()

                progress_label["text"] = "Hotovo."
            except Exception as e:
                messagebox.showerror("Chyba", str(e))
                progress_label["text"] = "Chyba pri spracovaní."

        progress_bar.pack(pady=10)
        progress_label.pack(pady=5)
        threading.Thread(target=spracuj).start()

    root = Tk()
    
     # Výpočet rozmerov a zarovnanie GUI do stredu obrazovky
    okno_sirka = 460
    okno_vyska = 300

    # Výpočet stredu obrazovky
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    x = int((screen_width / 2) - (okno_sirka / 2))
    y = int((screen_height / 2) - (okno_vyska / 2))

    root.geometry(f"{okno_sirka}x{okno_vyska}+{x}+{y}")

    # Nastavenie štýlu GUI (farby, progress bar)
    style = ttk.Style()
    style.theme_use("default")
    style.configure("green.Horizontal.TProgressbar", troughcolor="#C5D1A3", background="#6E7F46", thickness=20)

    root.title("Výber režimu objednávania")
    root.iconbitmap(extract_icon_temp())
    root.configure(bg="#C5D1A3")

    # Úvodné okno s výberom režimu
    vyber_rezim_frame = Frame(root, bg="#C5D1A3")
    vyber_rezim_frame.pack(pady=30)
    Label(vyber_rezim_frame, text="🦖 Zvoľ typ objednávania:", bg="#C5D1A3", font=("Segoe UI", 11, "bold")).pack(pady=10)
    Button(vyber_rezim_frame, text="✅ Vykrytie objednávok", command=zobraz_formular_old,
           bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4).pack(pady=5)
    Button(vyber_rezim_frame, text="📦 Objednávanie na sklad", command=zobraz_formular_new,
           bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4).pack(pady=5)
    
    # Tlačidlo späť na výber režimu
    spatne_tlacidlo = Button(root, text="← Späť na výber režimu",
    command=lambda: (
        formular_old.pack_forget(),
        formular_new.pack_forget(),
        spatne_tlacidlo.pack_forget(),
        progress_bar.pack_forget(),
        progress_label.pack_forget(),
        vyber_rezim_frame.pack(pady=30)
        ),
        bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4)

    # Formular pre starý režim
    formular_old = Frame(root, bg="#C5D1A3")
    zahrnut_var = BooleanVar()
    Checkbutton(formular_old, text="Zahrnúť bestsellery", variable=zahrnut_var, bg="#C5D1A3").pack()
    Button(formular_old, text="Vybrať Excel súbor",
           command=lambda: vyber_subor(zahrnut_bestsellery=zahrnut_var.get()),
           bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4).pack(pady=5)

    # Formular pre nový režim
    formular_new = Frame(root, bg="#C5D1A3")
    koef_b_var = StringVar()
    koef_n_var = StringVar()
    Label(formular_new, text="Koeficient pre Bestseller:", bg="#C5D1A3").pack()
    Entry(formular_new, textvariable=koef_b_var).pack()
    Label(formular_new, text="Koeficient pre ostatné produkty:", bg="#C5D1A3").pack()
    Entry(formular_new, textvariable=koef_n_var).pack()
    Button(formular_new, text="Vybrať Excel súbor",
           command=lambda: vyber_subor(
               koef_b=float(koef_b_var.get().replace(",", ".")),
               koef_n=float(koef_n_var.get().replace(",", "."))
           ),
           bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4).pack(pady=5)
    
    # Progress bar a status label
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=360, mode="determinate",
                                   style="green.Horizontal.TProgressbar")
    progress_label = Label(root, text="", bg="#C5D1A3")
    root.mainloop()
    
# Spustenie aplikácie pri spustení súboru
if __name__ == "__main__":
    spust_gui()
