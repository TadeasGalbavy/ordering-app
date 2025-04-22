import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, messagebox, ttk, Button, Label, Entry, StringVar, Checkbutton, BooleanVar, Frame, Toplevel
import os
import time
import threading
import shutil
import tempfile
import sys
from decimal import Decimal, getcontext

getcontext().prec = 6

def extract_icon_temp():
    if hasattr(sys, "_MEIPASS"):
        zdroj = os.path.join(sys._MEIPASS, "logo.ico")
        ciel = os.path.join(tempfile.gettempdir(), "temp_logo.ico")
        shutil.copyfile(zdroj, ciel)
        return ciel
    else:
        return "logo.ico"

def najdi_min_objednavku(skladom, objednavky, priemer, ciel_koef):
    skladom = Decimal(skladom)
    objednavky = Decimal(objednavky)
    priemer = Decimal(priemer) if priemer != 0 else Decimal("0.1")
    ciel_koef = Decimal(str(ciel_koef))

    for objednat in range(0, 100):
        aktualna_zasoba = skladom - objednavky + objednat
        aktualny_koef = aktualna_zasoba / priemer
        if round(aktualny_koef, 2) >= round(ciel_koef, 2):
            return objednat
    return 100

def vypocitaj_objednavku(row, koef_bestseller, koef_bezny):
    skladom = row["Reálne skladom"]
    objednavky = row["Počet nevybavených objednávok"] + row["Počet neštandardných objednávok"]
    priemer = row["Štvrťročný priemer"] if row["Štvrťročný priemer"] != 0 else 0.1

    if row["Dopredaj"] == "Ano" or row["Na objednávku"] == "Ano":
        return max(objednavky - skladom, 0)

    koef = koef_bestseller if row["Bestseller"] == "Ano" else koef_bezny
    return najdi_min_objednavku(skladom, objednavky, priemer, koef)

def uprav_excel_old(subor_cesta, vystup_cesta, zahrnut_bestsellery, progress_callback=None):
    df = pd.read_excel(subor_cesta)
    if progress_callback: progress_callback(10)
    df["Štvrťročný priemer"] = df["Štvrťročný priemer"].replace([0, '0'], 0.1)

    if zahrnut_bestsellery:
        df["Calculation"] = df.apply(lambda row: max(
            0,
            (2 - row["Reálne skladom"]) if row["Bestseller"] == "Ano" else
            row["Počet nevybavených objednávok"] + row["Počet neštandardných objednávok"] - row["Reálne skladom"]
        ), axis=1)
    else:
        df["Calculation"] = df["Reálne skladom"] - (
            df["Počet nevybavených objednávok"] + df["Počet neštandardných objednávok"])
        df["Calculation"] = df["Calculation"].apply(lambda x: 0 if x >= 0 else x * -1)

    if progress_callback: progress_callback(30)
    df.to_excel(vystup_cesta, index=False)
    wb = load_workbook(vystup_cesta)
    ws = wb.active
    podfarbi_a_oramuj(ws)
    wb.save(vystup_cesta)
    if progress_callback: progress_callback(100)
    return "OK"

def uprav_excel_new(subor_cesta, vystup_cesta, koef_bestseller, koef_bezny, progress_callback=None):
    df = pd.read_excel(subor_cesta)
    if progress_callback: progress_callback(10)
    df["Štvrťročný priemer"] = df["Štvrťročný priemer"].replace([0, '0'], 0.1)
    df["Calculation"] = df.apply(lambda row: vypocitaj_objednavku(row, koef_bestseller, koef_bezny), axis=1)

    if progress_callback: progress_callback(30)
    df.to_excel(vystup_cesta, index=False)
    wb = load_workbook(vystup_cesta)
    ws = wb.active

    skladom_col = objednavky1_col = objednavky2_col = priemer_col = calc_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Reálne skladom": skladom_col = get_column_letter(idx)
        elif cell.value == "Počet nevybavených objednávok": objednavky1_col = get_column_letter(idx)
        elif cell.value == "Počet neštandardných objednávok": objednavky2_col = get_column_letter(idx)
        elif cell.value == "Štvrťročný priemer": priemer_col = get_column_letter(idx)
        elif cell.value == "Calculation": calc_col = get_column_letter(idx)

    if skladom_col and objednavky1_col and objednavky2_col and priemer_col and calc_col:
        koef_col_idx = ws.max_column + 1
        ws.cell(row=1, column=koef_col_idx, value="Vypočítaný KOEF")
        for row in range(2, ws.max_row + 1):
            formula = f"=({skladom_col}{row}-{objednavky1_col}{row}-{objednavky2_col}{row}+{calc_col}{row})/{priemer_col}{row}"
            cell = ws.cell(row=row, column=koef_col_idx, value=formula)
            cell.number_format = "0.0"

    podfarbi_a_oramuj(ws)
    wb.save(vystup_cesta)
    if progress_callback: progress_callback(100)
    return "OK"

def podfarbi_a_oramuj(ws):
    fills = {
        "bledomodrá": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "olivová": PatternFill(start_color="808000", end_color="808000", fill_type="solid"),
        "oranžová": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "červená": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
    }
    farby_stlpcov = {
        "Ročný priemer": fills["bledomodrá"],
        "Polročný priemer": fills["bledomodrá"],
        "Štvrťročný priemer": fills["bledomodrá"],
        "Reálne skladom": fills["olivová"],
        "Počet nevybavených objednávok": fills["oranžová"],
        "Počet neštandardných objednávok": fills["oranžová"],
        "Calculation": fills["červená"],
    }
    for col in ws.iter_cols(min_row=1):
        header = col[0].column_letter + '1'
        stlpec_nazov = ws[header].value
        if stlpec_nazov in farby_stlpcov:
            for cell in col:
                if cell.value not in (None, '', 'NaN'):
                    cell.fill = farby_stlpcov[stlpec_nazov]

    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

def spust_gui():
    def zobraz_formular_old():
        vyber_rezim_frame.pack_forget()
        formular_old.pack(pady=10)
        spatne_tlacidlo.pack(pady=10)

    def zobraz_formular_new():
        vyber_rezim_frame.pack_forget()
        formular_new.pack(pady=10)
        spatne_tlacidlo.pack(pady=10)

    def vyber_subor(zahrnut_bestsellery=None, koef_b=None, koef_n=None):
        subor = filedialog.askopenfilename(title="Vyber Excel súbor", filetypes=[("Excel súbory", "*.xlsx")])
        if not subor:
            return

        progress_bar["value"] = 0
        progress_label["text"] = "Spracovanie prebieha..."

        def update_progress(value):
            progress_bar["value"] = value
            root.update_idletasks()
            time.sleep(0.15)

        def spracuj():
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    vystup = tmp.name
                if zahrnut_bestsellery is not None:
                    uprav_excel_old(subor, vystup, zahrnut_bestsellery, update_progress)
                else:
                    uprav_excel_new(subor, vystup, koef_b, koef_n, update_progress)

                
                for widget in root.winfo_children():
                    widget.pack_forget()

                Label(root, text="✅ Hotovo! Chceš uložiť výstupný súbor?", bg="#C5D1A3", font=("Segoe UI", 11)).pack(pady=15)

                def uloz():
                    nova_cesta = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
                    if nova_cesta:
                        shutil.copyfile(vystup, nova_cesta)
                        messagebox.showinfo("Uložené", f"Súbor bol uložený ako:\n{nova_cesta}")
                    root.destroy()

                Button(root, text="Uložiť ako...", command=uloz,
                    bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=10).pack(pady=5)

                Button(root, text="Zatvoriť", command=root.destroy,
                    bg="#999999", fg="white", font=("Segoe UI", 9), padx=8).pack()

                progress_label["text"] = "Hotovo."
            except Exception as e:
                messagebox.showerror("Chyba", str(e))
                progress_label["text"] = "Chyba pri spracovaní."

        progress_bar.pack(pady=10)
        progress_label.pack(pady=5)
        threading.Thread(target=spracuj).start()

    root = Tk()

    style = ttk.Style()
    style.theme_use("default")
    style.configure("green.Horizontal.TProgressbar", troughcolor="#C5D1A3", background="#6E7F46", thickness=20)

    root.title("Výber režimu objednávania")
    root.iconbitmap(extract_icon_temp())
    root.configure(bg="#C5D1A3")
    root.geometry("460x300")

    vyber_rezim_frame = Frame(root, bg="#C5D1A3")
    vyber_rezim_frame.pack(pady=30)
    Label(vyber_rezim_frame, text="Zvoľ typ objednávania:", bg="#C5D1A3", font=("Segoe UI", 11, "bold")).pack(pady=10)
    Button(vyber_rezim_frame, text="✅ Vykrytie objednávok", command=zobraz_formular_old,
           bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4).pack(pady=5)
    Button(vyber_rezim_frame, text="📦 Objednávanie na sklad", command=zobraz_formular_new,
           bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4).pack(pady=5)
    
    spatne_tlacidlo = Button(root, text="← Späť na výber režimu",
    command=lambda: (
        formular_old.pack_forget(),
        formular_new.pack_forget(),
        spatne_tlacidlo.pack_forget(),
        progress_bar.pack_forget(),
        progress_label.pack_forget(),
        vyber_rezim_frame.pack(pady=30)
        ),
        bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4)

    # Formular pre starý režim
    formular_old = Frame(root, bg="#C5D1A3")
    zahrnut_var = BooleanVar()
    Checkbutton(formular_old, text="Zahrnúť bestsellery", variable=zahrnut_var, bg="#C5D1A3").pack()
    Button(formular_old, text="Vybrať Excel súbor",
           command=lambda: vyber_subor(zahrnut_bestsellery=zahrnut_var.get()),
           bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4).pack(pady=5)

    # Formular pre nový režim
    formular_new = Frame(root, bg="#C5D1A3")
    koef_b_var = StringVar()
    koef_n_var = StringVar()
    Label(formular_new, text="Koeficient pre Bestseller (Áno):", bg="#C5D1A3").pack()
    Entry(formular_new, textvariable=koef_b_var).pack()
    Label(formular_new, text="Koeficient pre ostatné produkty:", bg="#C5D1A3").pack()
    Entry(formular_new, textvariable=koef_n_var).pack()
    Button(formular_new, text="Vybrať Excel súbor",
           command=lambda: vyber_subor(
               koef_b=float(koef_b_var.get().replace(",", ".")),
               koef_n=float(koef_n_var.get().replace(",", "."))
           ),
           bg="#6E7F46", fg="white", font=("Segoe UI", 10, "bold"), padx=8, pady=4).pack(pady=5)

    progress_bar = ttk.Progressbar(root, orient="horizontal", length=360, mode="determinate",
                                   style="green.Horizontal.TProgressbar")
    progress_label = Label(root, text="", bg="#C5D1A3")
    ttk.Style().configure("green.Horizontal.TProgressbar", troughcolor="#C5D1A3", background="#6E7F46", thickness=20)

    root.mainloop()

if __name__ == "__main__":
    spust_gui()